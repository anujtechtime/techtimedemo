from base64 import b64decode
from collections import OrderedDict
from io import BytesIO
from logging import getLogger

import base64


from openpyxl.styles import Alignment, Border, Side, Font

from openpyxl.drawing.image import Image

from PIL import Image as PILImage
from docx import Document
from docxcompose.composer import Composer
from docxtpl import DocxTemplate
import openpyxl
from jinja2 import Environment as Jinja2Environment
from requests import codes as codes_request, post as post_request
from requests.exceptions import RequestException

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, UserError
from odoo.http import request
from odoo.tools.safe_eval import safe_eval
import base64
import time
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter




try:
    from odoo.addons.gotenberg.service.utils import (
        get_auth,  # noqa
        convert_pdf_from_office_url,  # noqa
        check_gotenberg_installed,  # noqa
    )

    gotenberg_imported = True
except ImportError:
    gotenberg_imported = False

_logger = getLogger(__name__)


class IrActionsReport(models.Model):
    _inherit = "ir.actions.report"

    report_name = fields.Char(
        compute="_compute_report_name",
        inverse="_inverse_report_name",
        store=True,
        required=False,
    )
    report_type = fields.Selection(
        selection_add=[("docx-docx", "DOCX"), ("docx-pdf", "DOCX(PDF)"), ("docx-xlsx", "XlSX")],
    )
    report_docx_template = fields.Binary(
        string="Report docx template",
    )

    @api.depends("report_type", "model")
    def _compute_report_name(self):
        for record in self:
            if (
                record.report_type in ["docx-docx", "docx-pdf", "docx-xlsx"]
                and record.model
                and record.id
            ):
                record.report_name = "%s-docx_report+%s" % (record.model, record.id)
            else:
                record.report_name = False

    def _inverse_report_name(self):
        """TODO: write this method"""
        pass

    def retrieve_attachment(self, record):
        """
        Поиск существующего файла отчета во вложениях записи по:
        1. name
        2. res_model
        3. res_id
        """
        result = super().retrieve_attachment(record)
        if result:
            if self.report_type == "docx-docx":
                result = (
                    result.filtered(
                        lambda r: r.mimetype
                        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    )
                    or None
                )

            elif self.report_type == "docx-xlsx":
                result = (
                    result.filtered(
                        lambda r: r.mimetype
                        == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                    )
                    or None
                )
                    
            elif self.report_type == "docx-pdf":
                result = (
                    result.filtered(lambda r: r.mimetype == "application/pdf") or None
                )

        return result

    @api.model
    def _render_docx_pdf(self, res_ids=None, data=None):
        """
        Подготавливает данные для рендера файла отчета, вызывает метод рендера
        И обрабатывает результат рендера
        """
        if not data:
            data = {}
        data.setdefault("report_type", "pdf")

        # access the report details with sudo() but evaluation context as current user
        self_sudo = self.sudo()

        save_in_attachment = OrderedDict()
        # Maps the streams in `save_in_attachment` back to the records they came from
        stream_record = dict()
        if res_ids:
            Model = self.env[self_sudo.model]
            record_ids = Model.browse(res_ids)
            docx_record_ids = Model
            if self_sudo.attachment:
                for record_id in record_ids:
                    attachment = self_sudo.retrieve_attachment(record_id)
                    if attachment:
                        stream = self_sudo._retrieve_stream_from_attachment(attachment)
                        save_in_attachment[record_id.id] = stream
                        stream_record[stream] = record_id
                    if not self_sudo.attachment_use or not attachment:
                        docx_record_ids += record_id
            else:
                docx_record_ids = record_ids
            res_ids = docx_record_ids.ids

        if save_in_attachment and not res_ids:
            _logger.info("The PDF report has been generated from attachments.")
            self._raise_on_unreadable_pdfs(save_in_attachment.values(), stream_record)
            return self_sudo._post_pdf(save_in_attachment), "pdf"

        docx_content = self._render_docx(res_ids, data=data)

        pdf_content = (
            self._get_pdf_from_office(docx_content)
            if gotenberg_imported and check_gotenberg_installed()
            else None
        )

        if not pdf_content:
            raise UserError(
                _(
                    "Gotenberg converting service not available. The PDF can not be created."
                )
            )

        if res_ids:
            self._raise_on_unreadable_pdfs(save_in_attachment.values(), stream_record)
            _logger.info(
                "The PDF report has been generated for model: %s, records %s."
                % (self_sudo.model, str(res_ids))
            )
            return (
                self_sudo._post_pdf(
                    save_in_attachment, pdf_content=pdf_content, res_ids=res_ids
                ),
                "pdf",
            )
        return pdf_content, "pdf"

    @api.model
    def _render_docx_docx(self, res_ids=None, data=None):
        """
        Подготавливает данные для рендера файла отчета, вызывает метод рендера
        И обрабатывает результат рендера
        """
        if not data:
            data = {}
        data.setdefault("report_type", "xlsx")

        # access the report details with sudo() but evaluation context as current user
        self_sudo = self.sudo()

        save_in_attachment = OrderedDict()
        # Maps the streams in `save_in_attachment` back to the records they came from
        stream_record = dict()
        if res_ids:
            Model = self.env[self_sudo.model]
            record_ids = Model.browse(res_ids)
            docx_record_ids = Model
            if self_sudo.attachment:
                for record_id in record_ids:
                    attachment = self_sudo.retrieve_attachment(record_id)
                    if attachment:
                        stream = self_sudo._retrieve_stream_from_attachment(attachment)
                        save_in_attachment[record_id.id] = stream
                        stream_record[stream] = record_id
                    if not self_sudo.attachment_use or not attachment:
                        docx_record_ids += record_id
            else:
                docx_record_ids = record_ids
            res_ids = docx_record_ids.ids

        if save_in_attachment and not res_ids:
            _logger.info("The DOCS report has been generated from attachments.")
            return self_sudo._post_docx(save_in_attachment), "xlsx"

        docx_content = self._render_docx(res_ids, data=data)

        if res_ids:
            _logger.info(
                "The DOCS report has been generated for model: %s, records %s."
                % (self_sudo.model, str(res_ids))
            )
            _logger.info(
                "The DOCS report h@@@@@@@@@@@@@@@@@@@@@@2: %s"
                % str(docx_content)
            )
            return (
                self_sudo._post_docx(
                    save_in_attachment, docx_content=docx_content, res_ids=res_ids
                ),
                "xlsx",
            )


        return docx_content, "xlsx"

    @api.model
    def _render_docx_xlsx(self, res_ids=None, data=None):
        """
        Подготавливает данные для рендера файла отчета, вызывает метод рендера
        И обрабатывает результат рендера
        """
        if not data:
            data = {}
        data.setdefault("report_type", "xlsx")

        # access the report details with sudo() but evaluation context as current user
        self_sudo = self.sudo()

        save_in_attachment = OrderedDict()
        # Maps the streams in `save_in_attachment` back to the records they came from
        stream_record = dict()
        if res_ids:
            Model = self.env[self_sudo.model]
            record_ids = Model.browse(res_ids)
            docx_record_ids = Model
            if self_sudo.attachment:
                for record_id in record_ids:
                    attachment = self_sudo.retrieve_attachment(record_id)
                    if attachment:
                        stream = self_sudo._retrieve_stream_from_attachment(attachment)
                        save_in_attachment[record_id.id] = stream
                        stream_record[stream] = record_id
                    if not self_sudo.attachment_use or not attachment:
                        docx_record_ids += record_id
            else:
                docx_record_ids = record_ids
            res_ids = docx_record_ids.ids

        if save_in_attachment and not res_ids:
            _logger.info("The DOCS report has been generated from attachments.")
            return self_sudo._post_docx(save_in_attachment), "xlsx"

        docx_content = self._render_docx_x(res_ids, data=data)

        if res_ids:
            _logger.info(
                "The DOCS report has been generated for model: %s, records %s."
                % (self_sudo.model, str(res_ids))
            )
            _logger.info(
                "The DOCS report h@@@@@@@@@@@@@@@@@@@@@@2: %s"
                % str(docx_content)
            )
            return (
                self_sudo._post_docx(
                    save_in_attachment, docx_content=docx_content, res_ids=res_ids
                ),
                "xlsx",
            )


        return docx_content, "xlsx"    

    def _post_docx(self, save_in_attachment, docx_content=None, res_ids=None):
        """
        Добавляет сгенерированный файл в аттачменты
        """

        def close_streams(streams):
            for stream in streams:
                try:
                    stream.close()
                except Exception:
                    pass

        if len(save_in_attachment) == 1 and not docx_content:
            return list(save_in_attachment.values())[0].getvalue()

        streams = []

        if docx_content:
            # Build a record_map mapping id -> record
            record_map = {
                r.id: r
                for r in self.env[self.model].browse(
                    [res_id for res_id in res_ids if res_id]
                )
            }

            # _logger.info("self.model@@@@@@@@@@@@@@@@@@@@@@@.%s" % self.model)
            # new_stream = self._postprocess_docx_report(
            #     record_map[res_ids[0]], docx_content
            # )
            # If no value in attachment or no record specified, only append the whole docx.
            if not record_map or not self.attachment:
                streams.append(docx_content)
            else:
                if len(res_ids) == 1:
                    # Only one record, so postprocess directly and append the whole docx.
                    if (
                        res_ids[0] in record_map
                        and not res_ids[0] in save_in_attachment
                    ):
                        new_stream = self._postprocess_docx_report(
                            record_map[res_ids[0]], docx_content
                        )
                        # If the buffer has been modified, mark the old buffer to be closed as well.
                        if new_stream and new_stream != docx_content:
                            close_streams([docx_content])
                            docx_content = new_stream
                    streams.append(docx_content)
                else:
                    streams.append(docx_content)

        if self.attachment_use:
            for stream in save_in_attachment.values():
                streams.append(stream)

        if len(streams) == 1:
            result = streams[0].getvalue()
        else:
            try:
                result = self._merge_docx(streams)
            except Exception as e:
                _logger.exception(e)
                raise UserError(_("One of the documents, you try to merge is fallback"))

        close_streams(streams)
        return result

    def _postprocess_docx_report(self, record, buffer):
        """
        Непосредственно создает запись в ir.attachment
        """
        # attachment_name = safe_eval(self.attachment, {"object": record, "time": time})
        # if not attachment_name:
        #     return None
        
        _logger.info("bufferbuffer@@@@@@@@@@@@@@@@@@@@@@@.%s" % buffer)
        encoded_content = base64.b64encode(buffer.getvalue())
        attachment_vals = {
            "name": self.name,
            "datas": encoded_content,
            "res_model": self.model,
            "res_id": record.id,
            "type": "binary",
            'mimetype': "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        }


        
        try:
            attachment = self.env["ir.attachment"].create(attachment_vals)



            # Convert PDF to base64
            # pdf_base64 = base64.b64encode(pdf_content).decode('utf-8')
            remard_id.update({
                "attachment_file" : encoded_content,
                })
        except AccessError:
            _logger.info(
                "Cannot save DOCX report %r as attachment", attachment_vals["name"]
            )
        else:
            _logger.info(
                "The DOCX document %s is now saved in the database",
                attachment_vals["name"],
            )
        return buffer

    def _merge_docx(self, streams):
        """
        Объединяет несколько docx файлов в один
        """
        if streams:
            writer = Document(streams[0])
            composer = Composer(writer)
            for stream in streams[1:]:
                reader = Document(stream)
                composer.append(reader)
            return composer.getvalue()
        else:
            return streams

    def _render_docx(self, docids, data=None):
        """
        Получает данные для рендеринга и вызывает его.
        """
        if not data:
            data = {}
        data.setdefault("report_type", "docx")
        data = self._get_rendering_context(docids, data)
        return self._render_docx_template(self.report_docx_template, values=data)


    def _render_docx_x(self, docids, data=None):
        """
        Получает данные для рендеринга и вызывает его.
        """
        if not data:
            data = {}
        data.setdefault("report_type", "xlsx")
        data = self._get_rendering_context(docids, data)
        return self._render_docx_template_xlsx(self.report_docx_template, values=data)    
        
    def _render_docx_template(self, template, values=None):
        """
        Непосредственно рендеринг docx файла
        """
        if values is None:
            values = {}

        context = dict(self.env.context, inherit_branding=False)

        # Browse the user instead of using the sudo self.env.user
        user = self.env["res.users"].browse(self.env.uid)
        website = None
        if request and hasattr(request, "website"):
            if request.website is not None:
                website = request.website
                context = dict(
                    context,
                    translatable=context.get("lang")
                    != request.env["ir.http"]._get_default_lang().code,
                )

        values.update(
            time=time,
            context_timestamp=lambda t: fields.Datetime.context_timestamp(
                self.with_context(tz=user.tz), t
            ),
            user=user,
            res_company=user.company_id,
            website=website,
            web_base_url=self.env["ir.config_parameter"]
            .sudo()
            .get_param("web.base.url", default=""),
        )

        data = {key: value for key, value in values.items() if not callable(value)}
        functions = {key: value for key, value in values.items() if callable(value)}

        docx_content = BytesIO()
        jinja_env = Jinja2Environment()
        jinja_env.globals.update(**functions)

        with BytesIO(b64decode(template)) as template_file:
            doc = DocxTemplate(template_file)
            doc.render(data, jinja_env)
            doc.save(docx_content)

        docx_content.seek(0)

        return docx_content    

    def _render_docx_template_xlsx(self, template, values=None):
        """
        Непосредственно рендеринг docx файла
        """
        if values is None:
            values = {}

        context = dict(self.env.context, inherit_branding=False)

        # Browse the user instead of using the sudo self.env.user
        user = self.env["res.users"].browse(self.env.uid)
        website = None
        if request and hasattr(request, "website"):
            if request.website is not None:
                website = request.website
                context = dict(
                    context,
                    translatable=context.get("lang")
                    != request.env["ir.http"]._get_default_lang().code,
                )

        values.update(
            time=time,
            context_timestamp=lambda t: fields.Datetime.context_timestamp(
                self.with_context(tz=user.tz), t
            ),
            user=user,
            res_company=user.company_id,
            website=website,
            web_base_url=self.env["ir.config_parameter"]
            .sudo()
            .get_param("web.base.url", default=""),
        )

        data = {key: value for key, value in values.items() if not callable(value)}
        functions = {key: value for key, value in values.items() if callable(value)}



        docx_content = BytesIO()
        jinja_env = Jinja2Environment()
        jinja_env.globals.update(**functions)

        custumer_po_number = ""


        with BytesIO(b64decode(template)) as template_file:
            doc = openpyxl.load_workbook(template_file)
            ws = doc.active
            model_name = data.get("doc_model")
            if model_name == "sale.order":

                odoo_data = self.env['sale.order'].browse(int(data.get('doc_ids')[0]))

                rows_added = len(data)


                # shift_rows_down(ws, end_row, rows_added)

                

                docs = odoo_data

                if docs.custumer_req_number:
                    custumer_po_number = str(docs.custumer_req_number)

                placeholder_mapping = {
                    '{{docs.partner_id.display_name}}' : docs.partner_id.display_name,  # Replace with your actual field name
                    '{{docs.partner_id.state_id.name}}' : docs.partner_id.state_id.name,
                    '{{docs.partner_id.country_id.name}}' : docs.partner_id.country_id.name,
                    '{{docs.validity_date}}' : docs.validity_date,
                    '{{docs.currency_id.name}}' : docs.currency_id.name,
                    'QUOTATION #{{docs.display_name}}' : "QUOTATION #" +  docs.display_name,
                    '{{docs.date_order}}' : docs.date_order,
                    '{{number_inv}}' : docs.number_inv,
                    # 'Quotation Validity: {{docs.validity_date}}' : "Quotation Validity:" + docs.validity_date.date(),
                    "{{docs.amount_total}}" : docs.amount_total,
                    "RFQ: {{docs.custumer_po_number}}" : "RFQ: " + custumer_po_number
                    # "{{docs.order_line.product_id.display_name}}" : docs.order_line.product_id.display_name
                }


                # Iterate over cells and replace placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value in placeholder_mapping:
                            cell.value = placeholder_mapping[cell.value]

                        # Function to find the cell address of a specific value
                def find_marker(sheet, marker):
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value == marker:
                                return cell.coordinate
                    return None

                start_marker = '{{start}}'
                # end_marker = '{{end}}'

                start_cell = find_marker(ws, start_marker)

                print("start_cell@@@@@@@",start_cell)

                if not start_cell:
                    raise ValueError("Start or End marker not found in the template.")

                data_m = []
                count = 1
                for order in odoo_data.order_line:
                    # image_bytes = base64.b64decode(order.product_id.image_512)
                    # with open('image.png', 'wb') as image_file:
                    #     image_file.write(image_bytes)
                    data_m.append([
                        count,
                        order.product_id.display_name,
                        order.product_uom.display_name,
                        order.product_uom_qty,
                        str(order.price_unit),
                        order.price_subtotal,
                        order.product_id.image_512,
                    ])
                    count = count + 1
                print("data@@@@@@@@@@@@@@@@",data_m)    

                start_row = ws[start_cell].row

                def resize_image(image_path, max_width, max_height):
                    img = PILImage.open(image_path)
                    img.thumbnail((max_width, max_height), PILImage.LANCZOS)
                    resized_image_path = "resized_" + image_path
                    img.save(resized_image_path)
                    return resized_image_path

                def pixels_to_column_width(pixels):
                    return (pixels - 12) / 7 + 1

                # Function to convert pixels to Excel's row height units
                def pixels_to_row_height(pixels):
                    return pixels * 0.75

                start_col = 2

                # Define the table headers and data
                headers = ["Item No:", "Required Item", "Unit", "Qty", "Unit Price", "Amount", "Photo"]
                data = data_m + [
                    
                    ["", "SUBTOTAL", "", "", "", "$" + str(docs.amount_total), ""],
                    ["", "3.3% W/H Tax:", "", "", "", "", ""],
                    ["", "LOGISTICS FEE", "", "", "", "Included", ""],
                    ["", "TOTAL", "", "", "",  "$" + str(docs.amount_total), ""]
                ]

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                font_style = Font(size=12)


                def decode_and_save_image(image_base64, filename):
                    image_bytes = base64.b64decode(image_base64)
                    with open(filename, 'wb') as image_file:
                        image_file.write(image_bytes)

                for row_num, row_data in enumerate(data, start=start_row):
                    for col_num, cell_value in enumerate(row_data, start=start_col):
                        
                        if col_num != 8:
                            cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thin_border
                            cell.font = font_style 

                    if row_data[6]:
                        image_filename = f"image_{row_num}.png"
                        decode_and_save_image(row_data[6], image_filename)
                        resized_image_path = resize_image(image_filename, max_width=100, max_height=100)

                        # Insert the image into the worksheet
                        img = Image(resized_image_path)
                        img.anchor = f'H{row_num}'
                        ws.add_image(img)

                        column_width = pixels_to_column_width(100)
                        row_height = pixels_to_row_height(100)
                        
                        ws.column_dimensions[get_column_letter(col_num)].width = 100
                        ws.row_dimensions[row_num].height = 100


                # Merge cells for the specific format
                start_row = start_row + len(data_m) - 1
                ws.merge_cells(start_row=start_row + 1, start_column=start_col + 1, end_row=start_row + 1, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 2, start_column=start_col + 1, end_row=start_row + 2, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 3, start_column=start_col + 1, end_row=start_row + 3, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 4, start_column=start_col + 1, end_row=start_row + 4, end_column=start_col + 3)
                # ws.merge_cells(start_row=start_row + 4, start_column=start_col + 5, end_row=start_row + 4, end_column=start_col + 6)

                # Define the dimensions for the columns to accommodate text
                ws.column_dimensions[get_column_letter(start_col)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 40
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 3)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 4)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 5)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 6)].width = 15

                
                doc.save(docx_content)

            
            if model_name == "stock.picking":

                odoo_data = self.env['stock.picking'].browse(int(data.get('doc_ids')[0]))

                rows_added = len(data)


                # shift_rows_down(ws, end_row, rows_added)

                

                docs = odoo_data

                if docs.custumer_po_number:
                    custumer_po_number = str(docs.custumer_po_number)

                placeholder_mapping = {
                    '{{docs.partner_id.display_name}}' : docs.partner_id.display_name,  # Replace with your actual field name
                    '{{docs.partner_id.state_id.name}}' : docs.partner_id.state_id.name,
                    '{{docs.partner_id.country_id.name}}' : docs.partner_id.country_id.name,
                    'DELIVERY NOTE #{{docs.display_name}}' : "DELIVERY NOTE #" +  docs.display_name,
                    # '{{docs.date_order}}' : docs.date_order,
                    # 'Quotation Validity: {{docs.validity_date}}' : "Quotation Validity:" + docs.validity_date.date(),
                    "{{docs.scheduled_date.date()}}" : docs.scheduled_date.date(),
                    "PO No: {{docs.custumer_po_number}}" : "PO No: " + custumer_po_number
                    # "{{docs.order_line.product_id.display_name}}" : docs.order_line.product_id.display_name
                }


                # Iterate over cells and replace placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value in placeholder_mapping:
                            cell.value = placeholder_mapping[cell.value]

                        # Function to find the cell address of a specific value
                def find_marker(sheet, marker):
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value == marker:
                                return cell.coordinate
                    return None

                start_marker = '{{start}}'
                # end_marker = '{{end}}'

                start_cell = find_marker(ws, start_marker)

                print("start_cell@@@@@@@",start_cell)

                if not start_cell:
                    raise ValueError("Start or End marker not found in the template.")

                data_m = []
                count = 1
                for order in odoo_data.move_ids_without_package:
                    # image_bytes = base64.b64decode(order.product_id.image_512)
                    # with open('image.png', 'wb') as image_file:
                    #     image_file.write(image_bytes)
                    data_m.append([
                        count,
                        order.product_id.display_name,
                        order.product_uom.display_name,
                        order.product_uom_qty
                    ])
                    count = count + 1
                print("data@@@@@@@@@@@@@@@@",data_m)    

                start_row = ws[start_cell].row

                def resize_image(image_path, max_width, max_height):
                    img = PILImage.open(image_path)
                    img.thumbnail((max_width, max_height), PILImage.LANCZOS)
                    resized_image_path = "resized_" + image_path
                    img.save(resized_image_path)
                    return resized_image_path

                def pixels_to_column_width(pixels):
                    return (pixels - 12) / 7 + 1

                # Function to convert pixels to Excel's row height units
                def pixels_to_row_height(pixels):
                    return pixels * 0.75

                start_col = 2

                # Define the table headers and data
                headers = ["Item No:", "Required Item", "Unit", "Qty", "Unit Price", "Amount", "Photo"]
                data = data_m

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                font_style = Font(size=12)


                def decode_and_save_image(image_base64, filename):
                    image_bytes = base64.b64decode(image_base64)
                    with open(filename, 'wb') as image_file:
                        image_file.write(image_bytes)

                for row_num, row_data in enumerate(data, start=start_row):
                    for col_num, cell_value in enumerate(row_data, start=start_col):
                        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        cell.font = font_style 

                    # if row_data[6]:
                    #     image_filename = f"image_{row_num}.png"
                    #     decode_and_save_image(row_data[6], image_filename)
                    #     resized_image_path = resize_image(image_filename, max_width=100, max_height=100)

                    #     # Insert the image into the worksheet
                    #     img = Image(resized_image_path)
                    #     img.anchor = f'H{row_num}'
                    #     ws.add_image(img)

                    #     column_width = pixels_to_column_width(100)
                    #     row_height = pixels_to_row_height(100)
                        
                    #     ws.column_dimensions[get_column_letter(col_num)].width = 100
                    #     ws.row_dimensions[row_num].height = 100


                # Merge cells for the specific format
                start_row = start_row + len(data_m) - 1
                # ws.merge_cells(start_row=start_row + 1, start_column=start_col + 1, end_row=start_row + 1, end_column=start_col + 3)
                # ws.merge_cells(start_row=start_row + 2, start_column=start_col + 1, end_row=start_row + 2, end_column=start_col + 3)
                # ws.merge_cells(start_row=start_row + 3, start_column=start_col + 1, end_row=start_row + 3, end_column=start_col + 3)
                # ws.merge_cells(start_row=start_row + 4, start_column=start_col + 1, end_row=start_row + 4, end_column=start_col + 3)
                # # ws.merge_cells(start_row=start_row + 4, start_column=start_col + 5, end_row=start_row + 4, end_column=start_col + 6)

                # Define the dimensions for the columns to accommodate text
                ws.column_dimensions[get_column_letter(start_col)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 40
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 3)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 4)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 5)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 6)].width = 15

                
                doc.save(docx_content)  

            if model_name == "account.move":

                odoo_data = self.env['account.move'].browse(int(data.get('doc_ids')[0]))

                rows_added = len(data)


                # shift_rows_down(ws, end_row, rows_added)

                

                docs = odoo_data

                if docs.custumer_po_number:
                    custumer_po_number = str(docs.custumer_po_number)

                if docs.invoice_payment_term_id.id == 1:    
                    payment_method = "Cash"

                if docs.invoice_payment_term_id.id != 1:    
                    payment_method = "Payment Term"    


                placeholder_mapping = {
                    '{{docs.partner_id.display_name}}' : docs.partner_id.display_name,  # Replace with your actual field name
                    'state_id' : docs.partner_id.state_id.name,
                    'country_id' : docs.partner_id.country_id.name,
                    'Currency: {{docs.currency_id.name}}' : "Currency:" + docs.currency_id.name,
                    'INVOICE #{{docs.display_name}}' : "#" +  docs.display_name,
                    '{{docs.invoice_date}}' : docs.invoice_date,
                    # "RFQ: {{docs.custumer_po_number}}" : "RFQ: " +  custumer_po_number,
                    "PO No: {{docs.custumer_po_number}}" : "PO No: " + custumer_po_number,
                    "{{payment_method}}" : payment_method

                    # 'Quotation Validity: {{docs.validity_date}}' : "Quotation Validity:" + docs.validity_date.date(),
                    # "{{docs.order_line.product_id.display_name}}" : docs.order_line.product_id.display_name
                }


                # Iterate over cells and replace placeholders
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value in placeholder_mapping:
                            cell.value = placeholder_mapping[cell.value]

                        # Function to find the cell address of a specific value
                def find_marker(sheet, marker):
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value == marker:
                                return cell.coordinate
                    return None

                start_marker = '{{start}}'
                # end_marker = '{{end}}'

                start_cell = find_marker(ws, start_marker)

                print("start_cell@@@@@@@",start_cell)

                if not start_cell:
                    raise ValueError("Start or End marker not found in the template.")

                data_m = []
                count = 1
                for order in odoo_data.invoice_line_ids:
                    # image_bytes = base64.b64decode(order.product_id.image_512)
                    # with open('image.png', 'wb') as image_file:
                    #     image_file.write(image_bytes)
                    data_m.append([
                        count,
                        order.product_id.display_name,
                        order.product_uom_id.display_name,
                        order.quantity,
                        str(order.price_unit),
                        docs.currency_id.symbol +  str(order.price_subtotal),
                    ])
                    count = count + 1
                print("data@@@@@@@@@@@@@@@@",data_m)    

                start_row = ws[start_cell].row

                def resize_image(image_path, max_width, max_height):
                    img = PILImage.open(image_path)
                    img.thumbnail((max_width, max_height), PILImage.LANCZOS)
                    resized_image_path = "resized_" + image_path
                    img.save(resized_image_path)
                    return resized_image_path

                def pixels_to_column_width(pixels):
                    return (pixels - 12) / 7 + 1

                # Function to convert pixels to Excel's row height units
                def pixels_to_row_height(pixels):
                    return pixels * 0.75

                start_col = 2

                # Define the table headers and data
                headers = ["Item No:", "Required Item", "Unit", "Qty", "Unit Price", "Amount"]
                data = data_m + [
                    
                    ["", "SUBTOTAL", "", "", "", docs.currency_id.symbol + str(docs.amount_total)],
                    ["", "3.3% W/H Tax:", "", "", "", ""],
                    ["", "LOGISTICS FEE", "", "", "", "Included"],
                    ["", "TOTAL", "", "", "",  docs.currency_id.symbol + str(docs.amount_total)]
                ]

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                font_style = Font(size=12)


                def decode_and_save_image(image_base64, filename):
                    image_bytes = base64.b64decode(image_base64)
                    with open(filename, 'wb') as image_file:
                        image_file.write(image_bytes)

                for row_num, row_data in enumerate(data, start=start_row):
                    for col_num, cell_value in enumerate(row_data, start=start_col):
                        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                        cell.font = font_style 



                # Merge cells for the specific format
                start_row = start_row + len(data_m) - 1
                ws.merge_cells(start_row=start_row + 1, start_column=start_col + 1, end_row=start_row + 1, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 2, start_column=start_col + 1, end_row=start_row + 2, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 3, start_column=start_col + 1, end_row=start_row + 3, end_column=start_col + 3)
                ws.merge_cells(start_row=start_row + 4, start_column=start_col + 1, end_row=start_row + 4, end_column=start_col + 3)
                # ws.merge_cells(start_row=start_row + 4, start_column=start_col + 5, end_row=start_row + 4, end_column=start_col + 6)

                # Define the dimensions for the columns to accommodate text
                ws.column_dimensions[get_column_letter(start_col)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 1)].width = 40
                ws.column_dimensions[get_column_letter(start_col + 2)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 3)].width = 10
                ws.column_dimensions[get_column_letter(start_col + 4)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 5)].width = 15
                ws.column_dimensions[get_column_letter(start_col + 6)].width = 15

                
                doc.save(docx_content)        

        docx_content.seek(0)

        return docx_content

    def _get_pdf_from_office(self, content_stream):
        """
        Вызов конвертации docx в pdf с помощью gotenberg
        """
        result = None
        url = convert_pdf_from_office_url()
        auth = get_auth()
        try:
            response = post_request(
                url,
                files={"file": ("converted_file.xlsx", content_stream.read())},
                auth=auth,
            )
            if response.status_code == codes_request.ok:
                result = response.content
            else:
                _logger.warning(
                    "Gotenberg response: %s - %s"
                    % (response.status_code, response.content)
                )
        except RequestException as e:
            _logger.exception(e)
        finally:
            return result
